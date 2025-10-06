from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, date
import json
from bson import json_util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import tempfile
from decimal import Decimal
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# === DATA MODELS ===

class Product(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    sku: str
    description: Optional[str] = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ProductCreate(BaseModel):
    name: str
    sku: str
    description: Optional[str] = ""

class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    email: Optional[str] = ""
    address: str
    created_at: datetime = Field(default_factory=datetime.utcnow)

class CustomerCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    address: str

class InvoiceItem(BaseModel):
    product_id: str
    product_name: str
    sku: str
    quantity: int
    weight: float  # Actual weight from QR code
    purity: str = "18K"  # Selected purity for this item
    rate_per_gram: float  # Rate based on selected purity from gold rates
    amount: float
    labor_charges: float = 0.0  # Automatically calculated based on weight

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    customer_id: str
    customer_name: str
    customer_phone: str
    customer_address: str
    items: List[InvoiceItem]
    subtotal: float
    labor_charges: float = 0.0  # Total labor charges (sum of individual items)
    tax_included: bool = True  # Whether tax is included or excluded
    tax_percentage: float = 3.0
    tax_amount: float = 0.0
    discount_amount: float = 0.0  # Discount amount in rupees
    old_gold_value: float = 0.0  # Old gold value to be deducted
    old_silver_value: float = 0.0  # Old silver value to be deducted
    total_amount: float
    invoice_date: str  # Store as string to avoid BSON issues
    created_at: datetime = Field(default_factory=datetime.utcnow)

class InvoiceCreate(BaseModel):
    customer_id: str
    items: List[dict]  # {product_id, quantity, weight, labor_charges}
    tax_included: bool = False  # Default to without tax
    discount_amount: float = 0.0
    old_gold_value: float = 0.0
    old_silver_value: float = 0.0
    tax_percentage: float = 3.0  # Default GST for jewelry

class SalesRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_id: str
    product_id: str
    product_name: str
    sku: str
    quantity: int
    weight: float
    rate_per_gram: float
    amount: float
    labor_charges: float = 0.0
    sale_date: str  # Store as string to avoid BSON issues
    created_at: datetime = Field(default_factory=datetime.utcnow)

class GoldRate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    purity: str  # 18K, 20K, 22K, 24K, Silver
    rate_per_gram: float
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class PrintConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Page Settings
    pageSize: str = "A5"
    orientation: str = "landscape"
    margins: dict = {"top": 20, "bottom": 20, "left": 20, "right": 20}
    copiesPerPage: int = 2
    
    # Header Settings
    companyName: str = "HARI BABU SARRAF"
    companyAddress: str = "MOHALA CHOWK, PURANPUR"
    showLogo: bool = False
    logoPosition: str = "left"
    headerBackgroundColor: str = "#ffffff"
    headerTextColor: str = "#000000"
    companyNameFontSize: int = 14
    addressFontSize: int = 10
    
    # Invoice Title
    invoiceTitle: str = "ROUGH ESTIMATE"
    titleFontSize: int = 16
    titleColor: str = "#000000"
    titlePosition: str = "center"
    
    # Table Settings
    tableStyle: str = "modern"
    tableHeaderColor: str = "#333333"
    tableHeaderTextColor: str = "#ffffff"
    alternateRowColor: str = "#f5f5f5"
    tableBorderColor: str = "#cccccc"
    tableBorderWidth: int = 1
    cellPadding: int = 5
    
    # Table Columns
    columns: list = [
        {"name": "itemName", "label": "ITEM NAME", "width": 150, "show": True},
        {"name": "labor", "label": "LAB", "width": 80, "show": True},
        {"name": "weight", "label": "WEIGHT", "width": 80, "show": True},
        {"name": "amount", "label": "AMOUNT", "width": 100, "show": True}
    ]
    
    # Font Settings
    defaultFont: str = "Helvetica"
    tableFontSize: int = 8
    headerFontSize: int = 9
    
    # Totals Section
    totalsPosition: str = "right"
    totalsBackgroundColor: str = "#f9f9f9"
    finalTotalHighlight: bool = True
    finalTotalColor: str = "#2563eb"
    
    # Footer Settings
    showTerms: bool = True
    terms: str = "Terms & Conditions: All sales are final. Prices subject to change."
    showBankDetails: bool = True
    bankDetails: str = "Bank: State Bank of India | A/C: 1234567890 | IFSC: SBIN0001234"
    showContact: bool = True
    contactInfo: str = "Contact: 9690124010, 9456977703"
    footerFontSize: int = 7
    
    # Colors & Branding
    primaryColor: str = "#2563eb"
    secondaryColor: str = "#64748b" 
    accentColor: str = "#f59e0b"
    
    # Additional Elements
    showQRCode: bool = False
    showSignature: bool = True
    signatureText: str = "Authorized Signature"
    watermark: str = ""
    showPageNumbers: bool = False
    
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class GoldRateCreate(BaseModel):
    purity: str
    rate_per_gram: float

class GoldRateUpdate(BaseModel):
    rate_per_gram: float

# === PRODUCT APIS ===

@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate):
    product_dict = product.dict()
    product_obj = Product(**product_dict)
    await db.products.insert_one(product_obj.dict())
    return product_obj

@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find().to_list(1000)
    return [Product(**product) for product in products]

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str):
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return Product(**product)

@api_router.get("/products/by-sku/{sku}", response_model=Product)
async def get_product_by_sku(sku: str):
    product = await db.products.find_one({"sku": sku})
    if not product:
        raise HTTPException(status_code=404, detail=f"Product with SKU '{sku}' not found")
    return Product(**product)

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product: ProductCreate):
    existing_product = await db.products.find_one({"id": product_id})
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    updated_data = product.dict()
    await db.products.update_one({"id": product_id}, {"$set": updated_data})
    
    updated_product = await db.products.find_one({"id": product_id})
    return Product(**updated_product)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

# === CUSTOMER APIS ===

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate):
    customer_dict = customer.dict()
    customer_obj = Customer(**customer_dict)
    await db.customers.insert_one(customer_obj.dict())
    return customer_obj

@api_router.get("/customers", response_model=List[Customer])
async def get_customers():
    customers = await db.customers.find().to_list(1000)
    return [Customer(**customer) for customer in customers]

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return Customer(**customer)

# === INVOICE APIS ===

@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate):
    # Get customer details
    customer = await db.customers.find_one({"id": invoice_data.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get current gold rates for purity-based pricing
    gold_rates = {}
    gold_rates_cursor = db.gold_rates.find()
    async for rate in gold_rates_cursor:
        gold_rates[rate["purity"]] = rate["rate_per_gram"]
    
    # Process items and calculate totals
    invoice_items = []
    subtotal = 0.0
    total_labor_charges = 0.0
    
    for item_data in invoice_data.items:
        product = await db.products.find_one({"id": item_data["product_id"]})
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item_data['product_id']} not found")
        
        quantity = item_data["quantity"]
        weight = item_data["weight"]  # Weight from QR code/manual input
        purity = item_data.get("purity", "18K")  # Default to 18K if not specified
        
        # Get rate per gram based on selected purity from gold rates
        rate_per_gram = gold_rates.get(purity, 5500)  # Default rate if purity not found
        
        # Calculate amount based on weight and purity-based rate
        amount = weight * rate_per_gram
        
        # Automatic labor calculation based on weight rules:
        # Weight <= 5.000g: labor = ₹500
        # Weight > 5.000g: labor = weight × 100
        if weight <= 5.000:
            labor_charges = 500
        else:
            labor_charges = weight * 100
        
        invoice_item = InvoiceItem(
            product_id=product["id"],
            product_name=product["name"],
            sku=product["sku"],
            quantity=quantity,
            weight=weight,
            purity=purity,
            rate_per_gram=rate_per_gram,
            amount=amount,
            labor_charges=labor_charges
        )
        invoice_items.append(invoice_item)
        subtotal += amount
        total_labor_charges += labor_charges
    
    # Calculate with new formula: Subtotal + Labor + Tax - Discount - (Old Gold + Old Silver) = Final Total
    subtotal_with_labor = subtotal + total_labor_charges
    
    # Calculate taxes
    if invoice_data.tax_included:
        tax_amount = subtotal_with_labor * (3.0 / 100)  # Default 3% tax
        subtotal_with_tax = subtotal_with_labor + tax_amount
    else:
        tax_amount = 0.0
        subtotal_with_tax = subtotal_with_labor
    
    # Apply deductions: Subtotal + Tax - Discount - (Old Gold + Old Silver) = Final Total
    total_amount = subtotal_with_tax - invoice_data.discount_amount - invoice_data.old_gold_value - invoice_data.old_silver_value
    
    # Generate invoice number
    invoice_count = await db.invoices.count_documents({})
    invoice_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{invoice_count + 1:04d}"
    
    # Create invoice
    invoice = Invoice(
        invoice_number=invoice_number,
        customer_id=customer["id"],
        customer_name=customer["name"],
        customer_phone=customer["phone"],
        customer_address=customer["address"],
        items=invoice_items,
        subtotal=subtotal,
        labor_charges=total_labor_charges,
        tax_included=invoice_data.tax_included,
        tax_percentage=3.0,  # Default tax percentage
        tax_amount=tax_amount,
        discount_amount=invoice_data.discount_amount,
        old_gold_value=invoice_data.old_gold_value,
        old_silver_value=invoice_data.old_silver_value,
        total_amount=total_amount,
        invoice_date=date.today().isoformat()
    )
    
    await db.invoices.insert_one(invoice.dict())
    
    # Create sales records
    for item in invoice_items:
        sales_record = SalesRecord(
            invoice_id=invoice.id,
            product_id=item.product_id,
            product_name=item.product_name,
            sku=item.sku,
            quantity=item.quantity,
            weight=item.weight,
            rate_per_gram=item.rate_per_gram,
            amount=item.amount,
            labor_charges=item.labor_charges,
            sale_date=date.today().isoformat()
        )
        await db.sales_records.insert_one(sales_record.dict())
    
    return invoice

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices():
    invoices = await db.invoices.find().sort("created_at", -1).to_list(1000)
    return [Invoice(**invoice) for invoice in invoices]

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return Invoice(**invoice)

# === PDF GENERATION ===

async def create_invoice_pdf(invoice: Invoice) -> str:
    """Create PDF file for invoice in A5 landscape format with jewelry business layout (original + duplicate side-by-side)"""
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"Invoice_{invoice.invoice_number}.pdf")
    
    # Get print configuration
    try:
        print_config = await db.print_config.find_one({})
        if not print_config:
            # Use default config if none saved
            print_config = {
                'pageSize': 'A5',
                'orientation': 'landscape',
                'copiesPerPage': 2,
                'companyName': 'HARI BABU SARRAF',
                'companyAddress': 'MOHALA CHOWK, PURANPUR',
                'invoiceTitle': 'ROUGH ESTIMATE',
                'tableHeaderColor': '#333333',
                'tableBorderColor': '#000000',
                'showContact': True,
                'contactInfo': 'CONTACTS: 9690124010, 9456977703'
            }
    except Exception:
        # Default config if database error
        print_config = {
            'pageSize': 'A5',
            'orientation': 'landscape', 
            'copiesPerPage': 2,
            'companyName': 'HARI BABU SARRAF',
            'companyAddress': 'MOHALA CHOWK, PURANPUR',
            'invoiceTitle': 'ROUGH ESTIMATE',
            'tableHeaderColor': '#333333',
            'tableBorderColor': '#000000',
            'showContact': True,
            'contactInfo': 'CONTACTS: 9690124010, 9456977703'
        }
    
    # A5 landscape page size (210 x 148 mm = 595 x 420 points)
    from reportlab.lib.pagesizes import A5, landscape
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import black, white, HexColor
    
    # Create canvas for A5 landscape
    c = pdf_canvas.Canvas(file_path, pagesize=landscape(A5))
    width, height = landscape(A5)  # 595 x 420 points
    
    def draw_detailed_table(x_start, table_y, table_width, is_duplicate=False):
        """Draw detailed items table for both original and duplicate copies"""
        
        # Table headers
        col_widths = [table_width * 0.4, table_width * 0.2, table_width * 0.2, table_width * 0.2]
        col_x = [x_start + 5]
        for i in range(1, 4):
            col_x.append(col_x[-1] + col_widths[i-1])
        
        row_height = 12
        header_y = table_y
        
        # Header row
        c.setFillColor(HexColor('#333333'))
        c.rect(x_start + 5, header_y - row_height, table_width - 10, row_height, fill=1, stroke=1)
        
        c.setFillColor(white)
        header_font_size = 5 if is_duplicate else 6
        c.setFont("Helvetica-Bold", header_font_size)
        c.drawString(col_x[0] + 2, header_y - 8, "ITEM NAME")
        c.drawCentredString(col_x[1] + col_widths[1]/2, header_y - 8, "LAB")
        c.drawCentredString(col_x[2] + col_widths[2]/2, header_y - 8, "WEIGHT")
        c.drawCentredString(col_x[3] + col_widths[3]/2, header_y - 8, "AMOUNT")
        
        c.setFillColor(black)
        
        # Data rows
        current_y = header_y - row_height
        total_amount = 0
        total_weight = 0
        total_labor = 0
        
        for i, item in enumerate(invoice.items[:4]):  # Limit to 4 items for space
            # Alternating row colors
            if i % 2 == 1:
                c.setFillColor(HexColor('#f5f5f5'))
                c.rect(x_start + 5, current_y - row_height, table_width - 10, row_height, fill=1, stroke=0)
                c.setFillColor(black)
            
            # Draw borders
            c.rect(x_start + 5, current_y - row_height, table_width - 10, row_height, fill=0, stroke=1)
            
            data_font_size = 4 if is_duplicate else 5
            c.setFont("Helvetica", data_font_size)
            c.drawString(col_x[0] + 2, current_y - 8, item.product_name[:18 if is_duplicate else 20])
            c.drawCentredString(col_x[1] + col_widths[1]/2, current_y - 8, f"₹{item.labor_charges:.0f}")
            c.drawCentredString(col_x[2] + col_widths[2]/2, current_y - 8, f"{item.weight:.1f}g")
            c.drawCentredString(col_x[3] + col_widths[3]/2, current_y - 8, f"₹{item.amount:.0f}")
            
            total_amount += item.amount
            total_weight += item.weight
            total_labor += item.labor_charges
            current_y -= row_height
        
        # Total row
        c.setFillColor(HexColor('#333333'))
        c.rect(x_start + 5, current_y - row_height, table_width - 10, row_height, fill=1, stroke=1)
        
        c.setFillColor(white)
        total_font_size = 5 if is_duplicate else 6
        c.setFont("Helvetica-Bold", total_font_size)
        c.drawString(col_x[0] + 2, current_y - 8, "TOTAL")
        c.drawCentredString(col_x[1] + col_widths[1]/2, current_y - 8, f"₹{total_labor:.0f}")
        c.drawCentredString(col_x[2] + col_widths[2]/2, current_y - 8, f"{total_weight:.1f}g")
        c.drawCentredString(col_x[3] + col_widths[3]/2, current_y - 8, f"₹{total_amount:.0f}")
        
        c.setFillColor(black)
        
        # Totals section below table
        draw_totals_section(x_start, current_y - row_height - 10, table_width, is_duplicate)
    
    def draw_simplified_summary(x_start, table_y, table_width):
        """Draw simplified summary for duplicate copy"""
        c.setFont("Helvetica-Bold", 6)
        c.drawCentredString(x_start + table_width/2, table_y, "ITEM SUMMARY")
        
        summary_y = table_y - 15
        c.setFont("Helvetica", 5)
        
        for i, item in enumerate(invoice.items[:3]):  # Show max 3 items
            c.drawString(x_start + 8, summary_y - (i * 8), f"{item.product_name[:15]}: ₹{item.amount:.0f}")
        
        # Final total
        c.setFont("Helvetica-Bold", 6)
        c.drawCentredString(x_start + table_width/2, summary_y - 40, f"FINAL TOTAL: ₹{invoice.total_amount:.0f}")
    
    def draw_totals_section(x_start, totals_y, table_width, is_duplicate=False):
        """Draw jewelry business totals section"""
        
        # Calculate gold price per 10g (sample calculation)
        total_weight = sum(item.weight for item in invoice.items)
        gold_price_per_10g = 55000  # Default, should come from gold rates
        
        c.rect(x_start + 5, totals_y - 50, table_width - 10, 50, stroke=1, fill=0)
        
        totals_font_size = 4 if is_duplicate else 5
        c.setFont("Helvetica", totals_font_size)
        line_y = totals_y - 8
        
        c.drawString(x_start + 8, line_y, f"GOLD PRICE (22K/10g): ₹{gold_price_per_10g}")
        c.drawString(x_start + 8, line_y - 8, f"OLD GOLD: ₹{invoice.old_gold_value:.0f}")
        c.drawString(x_start + 8, line_y - 16, f"OLD SILVER: ₹{invoice.old_silver_value:.0f}")
        c.drawString(x_start + 8, line_y - 24, f"DISCOUNT: ₹{invoice.discount_amount:.0f}")
        c.drawString(x_start + 8, line_y - 32, f"Total Weight: {total_weight:.1f}g")
        
        # Final total
        final_total_font_size = 5 if is_duplicate else 6
        c.setFont("Helvetica-Bold", final_total_font_size)
        c.drawString(x_start + 8, line_y - 42, f"FINAL TOTAL: ₹{invoice.total_amount:.0f}")
    
    def draw_jewelry_invoice_copy(x_start, copy_width, copy_text, is_duplicate=False):
        """Draw a single jewelry invoice copy in the specified area"""
        
        # Header section
        header_y = height - 30
        
        # Copy indicator
        c.setFont("Helvetica-Bold", 8)
        c.drawCentredString(x_start + copy_width/2, header_y, copy_text)
        
        # Title
        c.setFont("Helvetica-Bold", 10 if not is_duplicate else 8)
        c.drawCentredString(x_start + copy_width/2, header_y - 15, print_config['invoiceTitle'])
        
        # Company name with underline
        c.setFont("Helvetica-Bold", 9 if not is_duplicate else 7)
        company_name = print_config['companyName']
        c.drawCentredString(x_start + copy_width/2, header_y - 28, company_name)
        
        # Add underline for company name
        company_width = c.stringWidth(company_name, "Helvetica-Bold", 9 if not is_duplicate else 7)
        c.line(x_start + copy_width/2 - company_width/2, header_y - 30, 
               x_start + copy_width/2 + company_width/2, header_y - 30)
        
        # Address
        c.setFont("Helvetica", 7 if not is_duplicate else 6)
        c.drawCentredString(x_start + copy_width/2, header_y - 40, print_config['companyAddress'])
        
        # Contact info
        if print_config.get('showContact'):
            c.drawCentredString(x_start + copy_width/2, header_y - 50, print_config.get('contactInfo', ''))
        
        # Customer details box
        customer_y = header_y - 70
        c.rect(x_start + 5, customer_y - 25, copy_width - 10, 25, stroke=1, fill=0)
        
        c.setFont("Helvetica", 6 if not is_duplicate else 5)
        c.drawString(x_start + 8, customer_y - 8, f"NAME: {invoice.customer_name[:25]}")
        c.drawString(x_start + 8, customer_y - 16, f"DATE: {invoice.invoice_date}")
        c.drawString(x_start + 8, customer_y - 24, f"INV NO.: {invoice.invoice_number}")
        
        # Both original and duplicate use the same detailed format
        draw_detailed_table(x_start, customer_y - 35, copy_width, is_duplicate)
        
        # Function content is handled by the new helper functions above
    
    # Draw side-by-side layout for A5 landscape
    copy_width = width / 2
    
    # Draw original copy (left side)
    draw_jewelry_invoice_copy(0, copy_width, "ORIGINAL", is_duplicate=False)
    
    # Draw vertical line separator
    c.line(copy_width, 30, copy_width, height - 30)
    
    # Draw duplicate copy (right side)
    draw_jewelry_invoice_copy(copy_width, copy_width, "DUPLICATE", is_duplicate=True)
    
    # Save the PDF
    c.save()
    return file_path

# === EXCEL GENERATION (KEPT FOR SALES REPORTS) ===

def create_invoice_excel(invoice: Invoice) -> str:
    """Create Excel file for invoice and return file path"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Styling
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=16)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "JEWELRY STORE INVOICE"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Invoice details
    ws['A3'] = "Invoice Number:"
    ws['B3'] = invoice.invoice_number
    ws['A4'] = "Date:"
    ws['B4'] = invoice.invoice_date
    
    # Customer details
    ws['D3'] = "Customer Name:"
    ws['E3'] = invoice.customer_name
    ws['D4'] = "Phone:"
    ws['E4'] = invoice.customer_phone
    ws['D5'] = "Address:"
    ws['E5'] = invoice.customer_address
    
    # Items table header
    headers = ["S.No", "Product Name", "SKU", "Qty", "Weight(g)", "Rate/g", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.border = border
    
    # Items data
    for row, item in enumerate(invoice.items, 8):
        ws.cell(row=row, column=1, value=row-7).border = border
        ws.cell(row=row, column=2, value=item.product_name).border = border
        ws.cell(row=row, column=3, value=item.sku).border = border
        ws.cell(row=row, column=4, value=item.quantity).border = border
        ws.cell(row=row, column=5, value=f"{item.weight:.2f}").border = border
        ws.cell(row=row, column=6, value=f"₹{item.rate_per_gram:.2f}").border = border
        ws.cell(row=row, column=7, value=f"₹{item.amount:.2f}").border = border
    
    # Totals
    last_row = len(invoice.items) + 8
    current_row = last_row
    
    ws[f'F{current_row}'] = "Subtotal:"
    ws[f'G{current_row}'] = f"₹{invoice.subtotal:.2f}"
    current_row += 1
    
    # Labor charges if any
    if invoice.labor_charges > 0:
        ws[f'F{current_row}'] = "Labor Charges:"
        ws[f'G{current_row}'] = f"₹{invoice.labor_charges:.2f}"
        current_row += 1
    
    # Tax if included
    if invoice.tax_included and invoice.tax_amount > 0:
        ws[f'F{current_row}'] = f"Tax ({invoice.tax_percentage}%):"
        ws[f'G{current_row}'] = f"₹{invoice.tax_amount:.2f}"
        current_row += 1
    
    ws[f'F{current_row}'] = "Total Amount:"
    ws[f'G{current_row}'] = f"₹{invoice.total_amount:.2f}"
    ws[f'F{current_row}'].font = header_font
    ws[f'G{current_row}'].font = header_font
    
    # Add note about tax
    if not invoice.tax_included:
        current_row += 2
        ws[f'A{current_row}'] = "Note: This invoice is without tax"
        ws[f'A{current_row}'].font = Font(italic=True)
    
    # Save file
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"Invoice_{invoice.invoice_number}.xlsx")
    wb.save(file_path)
    return file_path

def create_sales_excel(start_date: str, end_date: str, sales_records: List[dict]) -> str:
    """Create Excel file for sales report and return file path"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"SALES REPORT ({start_date} to {end_date})"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Date", "Invoice No", "Product Name", "SKU", "Qty", "Weight(g)", "Rate/g", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.border = border
    
    # Data
    total_amount = 0
    for row, record in enumerate(sales_records, 4):
        ws.cell(row=row, column=1, value=record["sale_date"]).border = border
        ws.cell(row=row, column=2, value=record["invoice_number"]).border = border
        ws.cell(row=row, column=3, value=record["product_name"]).border = border
        ws.cell(row=row, column=4, value=record["sku"]).border = border
        ws.cell(row=row, column=5, value=record["quantity"]).border = border
        ws.cell(row=row, column=6, value=f"{record['weight']:.2f}").border = border
        ws.cell(row=row, column=7, value=f"₹{record['rate_per_gram']:.2f}").border = border
        ws.cell(row=row, column=8, value=f"₹{record['amount']:.2f}").border = border
        total_amount += record["amount"]
    
    # Total
    last_row = len(sales_records) + 4
    ws[f'G{last_row}'] = "TOTAL:"
    ws[f'H{last_row}'] = f"₹{total_amount:.2f}"
    ws[f'G{last_row}'].font = header_font
    ws[f'H{last_row}'].font = header_font
    
    # Save file
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"Sales_Report_{start_date}_to_{end_date}.xlsx")
    wb.save(file_path)
    return file_path

@api_router.get("/invoices/{invoice_id}/download")
async def download_invoice(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    invoice_obj = Invoice(**invoice)
    file_path = await create_invoice_pdf(invoice_obj)
    
    return FileResponse(
        path=file_path,
        filename=f"Invoice_{invoice_obj.invoice_number}.pdf",
        media_type="application/pdf"
    )

@api_router.get("/invoices/{invoice_id}/print")
async def get_invoice_for_print(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return Invoice(**invoice)

@api_router.get("/sales/download")
async def download_sales_report(start_date: str, end_date: str):
    # Parse and validate dates
    try:
        start_parsed = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_parsed = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Convert to string for MongoDB query
        start = start_parsed.isoformat()
        end = end_parsed.isoformat()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Get sales records
    sales_pipeline = [
        {
            "$match": {
                "sale_date": {"$gte": start, "$lte": end}
            }
        },
        {
            "$lookup": {
                "from": "invoices",
                "localField": "invoice_id",
                "foreignField": "id",
                "as": "invoice_info"
            }
        },
        {
            "$unwind": "$invoice_info"
        }
    ]
    
    sales_cursor = db.sales_records.aggregate(sales_pipeline)
    sales_records = await sales_cursor.to_list(None)
    
    # Format data for Excel
    formatted_records = []
    for record in sales_records:
        formatted_records.append({
            "sale_date": record["sale_date"],
            "invoice_number": record["invoice_info"]["invoice_number"],
            "product_name": record["product_name"],
            "sku": record["sku"],
            "quantity": record["quantity"],
            "weight": record["weight"],
            "rate_per_gram": record["rate_per_gram"],
            "amount": record["amount"]
        })
    
    file_path = create_sales_excel(start_date, end_date, formatted_records)
    
    return FileResponse(
        path=file_path,
        filename=f"Sales_Report_{start_date}_to_{end_date}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === GOLD RATES APIS ===

@api_router.get("/gold-rates", response_model=List[GoldRate])
async def get_gold_rates():
    rates = await db.gold_rates.find().to_list(1000)
    return [GoldRate(**rate) for rate in rates]

@api_router.post("/gold-rates", response_model=GoldRate)
async def create_gold_rate(rate: GoldRateCreate):
    # Check if rate for this purity already exists
    existing_rate = await db.gold_rates.find_one({"purity": rate.purity})
    if existing_rate:
        raise HTTPException(status_code=400, detail=f"Rate for {rate.purity} already exists. Use update instead.")
    
    rate_dict = rate.dict()
    rate_obj = GoldRate(**rate_dict)
    await db.gold_rates.insert_one(rate_obj.dict())
    return rate_obj

@api_router.put("/gold-rates/{purity}", response_model=GoldRate)
async def update_gold_rate(purity: str, rate_update: GoldRateUpdate):
    existing_rate = await db.gold_rates.find_one({"purity": purity})
    if not existing_rate:
        raise HTTPException(status_code=404, detail=f"Rate for {purity} not found")
    
    updated_data = {
        "rate_per_gram": rate_update.rate_per_gram,
        "updated_at": datetime.utcnow()
    }
    
    await db.gold_rates.update_one({"purity": purity}, {"$set": updated_data})
    updated_rate = await db.gold_rates.find_one({"purity": purity})
    return GoldRate(**updated_rate)

@api_router.get("/gold-rates/{purity}", response_model=GoldRate)
async def get_gold_rate_by_purity(purity: str):
    rate = await db.gold_rates.find_one({"purity": purity})
    if not rate:
        raise HTTPException(status_code=404, detail=f"Rate for {purity} not found")
    return GoldRate(**rate)

# Initialize default gold rates if they don't exist
@api_router.post("/gold-rates/initialize")
async def initialize_default_rates():
    default_rates = [
        {"purity": "18K", "rate_per_gram": 4500.0},
        {"purity": "20K", "rate_per_gram": 5000.0},
        {"purity": "22K", "rate_per_gram": 5500.0},
        {"purity": "24K", "rate_per_gram": 6000.0},
        {"purity": "Silver", "rate_per_gram": 80.0}
    ]
    
    initialized_count = 0
    for rate_data in default_rates:
        existing = await db.gold_rates.find_one({"purity": rate_data["purity"]})
        if not existing:
            rate_obj = GoldRate(**rate_data)
            await db.gold_rates.insert_one(rate_obj.dict())
            initialized_count += 1
    
    return {"message": f"Initialized {initialized_count} default gold rates"}

# === DASHBOARD APIS ===

@api_router.post("/dashboard/reset-sales")
async def reset_sales_data():
    """Reset all sales data (invoices and sales records) while keeping products and customers"""
    try:
        # Delete all invoices
        invoices_result = await db.invoices.delete_many({})
        
        # Delete all sales records
        sales_result = await db.sales_records.delete_many({})
        
        # Reset stock quantities back to original values if needed
        # (Optional: You might want to keep current stock levels)
        
        return {
            "message": "Sales data reset successfully",
            "deleted_invoices": invoices_result.deleted_count,
            "deleted_sales_records": sales_result.deleted_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error resetting sales data: {str(e)}")

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    total_products = await db.products.count_documents({})
    total_customers = await db.customers.count_documents({})
    total_invoices = await db.invoices.count_documents({})
    
    # Today's sales
    today = date.today().isoformat()
    today_sales = await db.sales_records.aggregate([
        {"$match": {"sale_date": today}},
        {"$group": {"_id": None, "total_amount": {"$sum": "$amount"}}}
    ]).to_list(1)
    
    today_sales_amount = today_sales[0]["total_amount"] if today_sales else 0
    
    return {
        "total_products": total_products,
        "total_customers": total_customers,
        "total_invoices": total_invoices,
        "today_sales": today_sales_amount
    }

# === PRINT CONFIGURATION APIS ===

@api_router.get("/print-config", response_model=PrintConfig)
async def get_print_config():
    config = await db.print_config.find_one({})
    if not config:
        # Return default configuration
        default_config = PrintConfig()
        return default_config
    return PrintConfig(**config)

@api_router.post("/print-config", response_model=PrintConfig)
async def save_print_config(config: PrintConfig):
    # Remove existing config (we only store one)
    await db.print_config.delete_many({})
    
    # Save new config
    await db.print_config.insert_one(config.dict())
    return config

@api_router.put("/print-config", response_model=PrintConfig)
async def update_print_config(config: PrintConfig):
    # Update existing config or create new one
    existing_config = await db.print_config.find_one({})
    if existing_config:
        await db.print_config.update_one({"id": existing_config["id"]}, {"$set": config.dict()})
    else:
        await db.print_config.insert_one(config.dict())
    return config

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()